from lib2to3.pgen2.pgen import DFAState
import numpy as np
import pandas as pd 
import matplotlib.pyplot as plt
import pandas as pd
import os 
import sys 
import matplotlib 
plt.rcParams['figure.constrained_layout.use'] = True
matplotlib.rc('xtick', labelsize=10) 
matplotlib.rc('ytick', labelsize=10) 
matplotlib.rcParams.update({'font.size': 15})

def save_error(df_plot, res, par, test):
    from openpyxl import load_workbook
    path = "parameters_21feb_2.xlsx"
    book = load_workbook(path)
    writer = pd.ExcelWriter(path, engine = 'openpyxl')
    writer.book = book

    df_plot[ "Flex Power (kW)" ] = "err" 
    df_plot[ "ASAP Power (kW)" ] = "err" 
    df_plot[ "Energy Delivered, Flex (kWh)" ] = "err" 

    df_plot['tariff_flex (cents/hour)'] = "err"

    df_plot['tariff_asap (cents/hour)'] = "err"
    
    df_plot['z0'] = str(list(par.z0.reshape((4,))))
    df_plot['v0'] = str(list(par.v0.reshape((3,))))

    df_plot['prob_asap'] = "err"
    df_plot['prob_flex'] = "err"
    df_plot['prob_leave'] = "err"
    df_plot['e_need'] = res['e_need']
    
    df_plot.to_excel(writer,sheet_name="test_{}".format(test))

    df_plot['']
    writer.save()
    writer.close()

    

def save_the_data(df_plot, res, par, test):
    from openpyxl import load_workbook
    path = "par_22_feb.xlsx"
    book = load_workbook(path)
    writer = pd.ExcelWriter(path, engine = 'openpyxl')
    writer.book = book

    df_plot['tariff_flex (cents/hour)'] = res['tariff_flex'] * 6.6

    df_plot['tariff_asap (cents/hour)'] = res['tariff_asap'] * 6.6
    
    df_plot['z0'] = str(list(par.z0.reshape((4,))))
    df_plot['v0'] = str(list(par.v0.reshape((3,))))

    df_plot['prob_asap'] = res['prob_asap']
    df_plot['prob_flex'] = res['prob_flex']
    df_plot['prob_leave'] = res['prob_leave']
    df_plot['e_need'] = res['e_need']
    df_plot['res'] = res['J'][-1]
    
    df_plot.to_excel(writer,sheet_name="test_{}".format(test))

    writer.save()
    writer.close()


def record_results(event, par, prb, opt, res, test, vis = True):
    days = 0 
    num_interval= len(par.TOU)
    interval_size = par.Ts

    start_hour = int(res["time_start"] * par.Ts // 1)
    start_minute = int(np.ceil(60 * ( res['time_start'] *par.Ts % 1)))
    flex_end_hour = int(res["time_end_flex"] * par.Ts // 1)
    flex_end_minute= int(np.ceil(60 * ( res['time_end_flex'] *par.Ts % 1)))
    asap_end_hour = int(res["time_end_asap"] * par.Ts // 1)
    asap_end_minute = int(np.ceil(60 * ( res['time_end_asap'] *par.Ts % 1)))

    # 5 min interval
    timerange = pd.date_range("1/1/2021", periods=num_interval, freq="{}min".format(5))
    ## all time frame 
    TOU_cost = pd.Series(par.TOU, index = timerange, name = "TOU Cost ($)")


    # Event parameters: arrival_time, soc_init, soc_need, departure
    opt_time_start  = pd.Timestamp(year=2021, month=1, day=1, hour = start_hour, minute = start_minute)
    opt_time_end_flex  =pd.Timestamp(year=2021, month=1, day=1, hour = flex_end_hour,  minute = flex_end_minute)
    opt_time_end_asap  = pd.Timestamp(year=2021, month=1, day=1, hour = asap_end_hour, minute = asap_end_minute)

    flex_ts = pd.date_range(start=opt_time_start , end=opt_time_end_flex, freq="{}H".format(par.Ts))
    asap_ts = pd.date_range(start=opt_time_start , end=opt_time_end_asap, freq="{}H".format(par.Ts))

    # Optimization outputs 

    ## Question: Why flex output length mismatches the timelength? 
    # print(opt.opt_flex_powers.shape)
    opt_flex_powers = pd.Series(res['flex_powers'].reshape(len(res['flex_powers']),), index = flex_ts[:-1], name = "Flew Power (kW)")
    opt_asap_powers = pd.Series(res['asap_powers'].reshape(len(res['asap_powers']),), index = asap_ts[:-1], name = "ASAP Power (kW)")
    opt_flex_energy = pd.Series(res['flex_e_delivered'].reshape(len(res['flex_e_delivered']),), index = flex_ts, name = "Energy Delivered (kWh)")
    # event = prb.event
    # print(event)

    ## PLOT FOR 2 days 
    days =2 
    ts = pd.date_range(timerange[0], periods=num_interval * days, freq="{}H".format(interval_size ))
    df_plot = pd.DataFrame(index = ts, 
                            data = {"TOU Cost (cents/hour)": list(TOU_cost.values* 6.6) * days, 
                            "Flex Power (kW)": opt_flex_powers, 
                            "ASAP Power (kW)":opt_asap_powers, 
                            "Energy Delivered, Flex (kWh)":opt_flex_energy})

    df_plot["Energy Delivered, Flex (kWh)"] = df_plot["Energy Delivered, Flex (kWh)"].fillna(method='bfill').fillna(method='ffill')
    df_plot = df_plot.fillna(0)
    df_plot['Station Max Power (kW)'] = event['pow_max']
    df_plot['Station Min Power (kW)'] = event['pow_min']
    # df_plot['Arrival SOC'] = event['SOC_init']
    # df_plot['Departure SOC Demand'] = event['SOC_need']
    df_plot['Arrival Time'] = event['time']
    df_plot['Duration'] = event['duration']

    if vis == True:
        plot_power_energy( df_plot, prb, test, start_hour,start_minute,flex_end_hour ,flex_end_minute)
        plot_prices_and_probs(res, test)
        save_the_data(df_plot, res, par, test)
    else: 
        save_error(df_plot, res, par, test)



def plot_arrival_departure(ax_list,start_hour,start_minute,flex_end_hour ,flex_end_minute):
    for ax in ax_list:
        arrival_time = pd.Timestamp(year=2021, month=1, day=1, hour = start_hour, minute = start_minute)
        departure_time = pd.Timestamp(year=2021, month=1, day=1, hour = flex_end_hour, minute = flex_end_minute)
        ax.axvspan(arrival_time, departure_time, facecolor='b', alpha=0.25)

def plot_max_min_power(ax_list, df_plot):
    for ax in ax_list:
        ax.plot(df_plot['Station Max Power (kW)'], label="Station Max Power (kW", color = 'red',linestyle='--')
        # ax.plot(df_plot['Station Min Power (kW)'],label="Station Max Power (kW")))

def x_axis_labels(ax_list, cuts=[0,6,12,18]):
    import matplotlib.dates as mdates
    hlocator = mdates.HourLocator(byhour=cuts, interval=1, tz=None)
    for ax in ax_list:
        # ax.xaxis.set_major_locator(loc)
        ax.xaxis.set_major_locator(hlocator)
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%H"))

def add_legend(ax_list):
    for ax in ax_list:
        ax.legend()

def plot_power_energy(df_plot, prb, test, start_hour,start_minute,flex_end_hour ,flex_end_minute):
    plt.rcParams['figure.constrained_layout.use'] = True

    fig,(ax1, ax2, ax3, ax4) = plt.subplots(4,1,figsize=(12,9) ,)

    x_lim  = (list(df_plot.index)[0],list(df_plot.index)[-1])
    ax1.plot(df_plot['TOU Cost (cents/hour)'])

    ax1.set_xlim(x_lim)
    ax1.grid()
    ax1.set_xlabel("Time")
    ax1.set_ylabel("TOU (cents / hour)")


    ax2.plot(df_plot["Flex Power (kW)"], label = "Power Delivered (kW)")
    ax2.set_xlim(x_lim)
    ax2.grid()
    ax2.set_xlabel("Time")
    ax2.set_ylabel("Power (kW)")
    ax2.set_title("Flex Charging Optimal Power")

    ax3.plot(df_plot["ASAP Power (kW)"], label = "Power Delivered (kW)")
    ax3.set_xlim(ax1.set_xlim(x_lim))
    ax3.grid()
    ax3.set_xlabel("Time")
    ax3.set_ylabel("Power (kW)")
    ax3.set_title("ASAP Charging Optimal Power")

    ax4.plot(df_plot["Energy Delivered, Flex (kWh)"], label  = 'Energy Delivered (kWh)')
    ax4.axhline(prb.e_need,  linestyle='--',color='red', label = 'Energy Demand (kWh)')
    ax4.set_xlim(x_lim)
    ax4.grid()
    ax4.set_xlabel("Time")
    ax4.set_ylabel("Energy Level (kWh)")

    ax4.set_title("Flex Charging Energy Delivered (kWh)")

    plot_arrival_departure([ax2,ax3],  start_hour,start_minute,flex_end_hour ,flex_end_minute)
    plot_max_min_power([ax2,ax3], df_plot)
    x_axis_labels([ax1,ax2,ax3,ax4], cuts=[0,5,10,15,20])
    add_legend([ax1,ax2,ax3,ax4])
    # fig.suptitle("{} , Solver=Gurobi".format(objective))
    plt.savefig(os.path.join("Figures/", "test_full_code_{}.pdf".format(test)),format='pdf')

    plt.show()
    plt.close()

# PLOT FOR EACH ITERATION
def plot_prices_and_probs(res, test):
    fig,(axes) = plt.subplots(5,1,figsize=(12,12))
    # Objective Value for each iteration 
    ax = axes[0]
    ax.plot(res['J'], label = 'Total Objective Value')
    ax.plot(res['J_sub'][0],label='Flex Charging Objective')
    ax.plot(res['J_sub'][1],label='Asap Charging Objective')
    ax.plot(res['J_sub'][2],label='Leave Objective')
    ax.margins(0)
    ax.grid()
    ax.set_xlabel("Iteration")
    ax.set_ylabel("$J(x,z,v)$")
    ax.set_title("Objective Value")
    ax.legend()

    # Flex Charging Revenue for each iteration and flex charging costs 
    ax = axes[1]
    ax.plot(res['rev_flex'],linewidth = 2)
    ax.set_xlim(0,res['num_iter'] - 1)
    ax.grid()
    ax.set_xlabel("Iteration")
    ax.set_ylabel("cents")
    ax.set_title("Flex Charging Revenue")

    ax_twin  = ax.twinx()
    ax_twin.plot(res['z_iter'][0,:] * 6.6,linestyle = "--",color ="red",linewidth = 2)
    ax_twin.set_ylabel("cents / hour", color = 'red')


    # Asap Charging Revenue for each iteration and asap charging costs 
    ax = axes[2]
    ax.plot(res['rev_asap'],linewidth = 2)
    ax.set_xlim(0,res['num_iter'] - 1)
    ax.grid()
    ax.set_xlabel("Iteration")
    ax.set_ylabel("cents")
    ax.set_title("ASAP Charging Revenue")

    ax_twin  = ax.twinx()
    ax_twin.plot(res['z_iter'][1,:] * 6.6,linestyle = "--",color ="red",linewidth = 2)
    ax_twin.set_ylabel("cents / hour", color = 'red')

    # Probability of Charging Flex, Asap and Leaving at each iteration 
    ax=axes[3]
    ax.bar(x = np.arange(res['num_iter']),height = res['v_iter'][0,:], label = 'Flex Charging')
    ax.bar(x = np.arange(res['num_iter']),height = res['v_iter'][1,:], label = 'ASAP Charging', bottom = res['v_iter'][0,:])
    ax.bar(x = np.arange(res['num_iter']),height = res['v_iter'][2,:], label = 'Leave', bottom = res['v_iter'][1,:] + res['v_iter'][0,:])
    ax.legend()
    ax.grid()
    ax.set_xlabel("Iteration")
    ax.set_ylabel("$P(i)$")
    ax.set_title("Probability")

    # Price of each option maybe 
    ax=axes[4]
    ax.bar(x = np.arange(res['num_iter']),height = res['z_iter'][0,:] * 6.6, label = 'Flex Charging')
    ax.bar(x = np.arange(res['num_iter']),height = res['z_iter'][1,:] * 6.6, label = 'ASAP Charging', bottom = res['z_iter'][0,:]* 6.6)
    # ax.bar(x = np.arange(res['num_iter']),height = res['z_iter'][2,:], label = 'Overstay', bottom = res['z_iter'][0,:] + res['z_iter'][1,:])
    # ax.bar(x = np.arange(res['num_iter']),height = res['z_iter'][3,:], label = 'Leave',bottom = (res['z_iter'][0,:] + res['z_iter'][1,:] + res['z_iter'][2,:]))
    ax.legend()
    ax.grid()
    ax.set_xlabel("Iteration")
    ax.set_title("Price")
    ax.set_ylabel("cents / hour", color = 'red')
    plt.savefig(os.path.join("Figures/", "power_price_{}.pdf".format(test)),format='pdf')

    # 